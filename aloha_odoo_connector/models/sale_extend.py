# -*- coding: utf-8 -*-

from odoo import models, exceptions, fields, _

import xlrd
import xlwt
from xlwt import Workbook
import paramiko
import os
from datetime import date
from ftplib import FTP
import csv
import openpyxl
import xlsxwriter
import base64
import io
import datetime
from datetime import timedelta



class Sale(models.Model):
    _inherit = "sale.order"

    aloha_file_path = fields.Char("Aloha File Path",copy=False,help='File Path Of Aloha XLSX')
    worksheet_name = fields.Char("Worksheet Name",copy=False,help="Give a worksheet name")
    aloha_connection_done = fields.Boolean("Aloha Connection Done",default=False,copy=False)
    aloha_date = fields.Date("Aloha Date")

    def get_remote_csv(self):
        def make_directory(new_dir):
            os.mkdir(new_dir)
            os.chmod(new_dir, 0o777)
        dummy_aloha_file_path = dummy_aloha_inv_file_path = ''
        odoo_dir = os.path.join(os.environ['HOME'], "Aloha")
        if not os.path.exists(odoo_dir):
            make_directory(odoo_dir)
        xlsx_dir_path = os.path.join(odoo_dir, "XLSX")
        if not os.path.exists(xlsx_dir_path):
            make_directory(xlsx_dir_path)
        csv_dir_path = os.path.join(odoo_dir, "CSV")
        if not os.path.exists(csv_dir_path):
            make_directory(csv_dir_path)
        pre_date = datetime.datetime.today() - datetime.timedelta(days=1)
        today_date = pre_date.strftime("%Y%m%d")
        new_csv_file_name = today_date+'.csv'
        csv_data = []
        doc_ob = self.env['documents.document'].search([('name','=',new_csv_file_name)],limit=1)
        if doc_ob:
            val = base64.b64decode(doc_ob.datas)
            csv_file_path=csv_dir_path+'/'+today_date+'.xlsx'
            workbook = xlsxwriter.Workbook(csv_file_path)
            worksheet = workbook.add_worksheet()
            workbook.close()
            output_measurement_file = csv_dir_path+'/'+today_date+'.xlsx'
            with open(output_measurement_file, 'wb') as f:
                f.write(val)
                f.close()
                dummy_aloha_file_path = str(csv_file_path)
        inv_csv_name = today_date+'_inv'+'.csv'
        inv_csv_file = self.env['documents.document'].search([('name','=',inv_csv_name)],limit=1)
        if inv_csv_file:
            inv_val = base64.b64decode(inv_csv_file.datas)
            csv_inv_file_path=csv_dir_path+'/'+today_date+'_inv'+'.xlsx'
            workbook = xlsxwriter.Workbook(csv_inv_file_path)
            worksheet = workbook.add_worksheet()
            workbook.close()
            inv_dummy_file = csv_dir_path+'/'+today_date+'_inv'+'.xlsx'
            with open(inv_dummy_file, 'wb') as f:
                f.write(inv_val)
                f.close()
                dummy_aloha_inv_file_path = str(inv_dummy_file)
            return dummy_aloha_file_path,xlsx_dir_path,dummy_aloha_inv_file_path
        else:
            return None,None,None

    #Fetch data from xlsx and add line in current sale order
    def get_aloha_line(self):
        dummy_aloha_file_path,xlsx_dir_path,dummy_aloha_inv_file_path = self.get_remote_csv()
        sale_date = fields.Datetime.now()- timedelta(days=1)
        formated = sale_date.strftime("%m/%d/%Y %H:%M:%S")
        now = datetime.datetime.strptime(formated, "%m/%d/%Y %H:%M:%S")
        if dummy_aloha_file_path and xlsx_dir_path:
            new_order_id = self.env['sale.order'].create({'partner_id':3,'aloha_date':now})
            pre_date = datetime.datetime.today() - datetime.timedelta(days=1)
            today_date = pre_date.strftime("%Y%m%d")
            csv_data = []
            with open(dummy_aloha_file_path) as file_obj:
                reader = csv.reader(file_obj)
                for row in reader:
                    csv_data.append(row)
            real_xlsx_path = xlsx_dir_path+'/'+today_date+'.xlsx'
            workbook = xlsxwriter.Workbook(real_xlsx_path)
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            for row in csv_data:
                sheet.append(row)
            workbook.save(real_xlsx_path)
            new_order_id.aloha_file_path = real_xlsx_path
            workbook = xlrd.open_workbook(new_order_id.aloha_file_path)
            worksheet = workbook.sheet_by_name(workbook.sheet_names()[0])
            num_rows = worksheet.nrows
            curr_row = 12
            missed_list = []
            updated_num_rows = 0
            while curr_row < num_rows:
                row1 = worksheet.row(curr_row)
                try:
                    if row1[0].value == 'SUMMARY SECTION':
                        updated_num_rows = curr_row-8
                        break
                except IOError:
                    print("Formate Error")
                curr_row = curr_row + 1
            new_curr_row = 12
            order_line = []
            while new_curr_row < updated_num_rows:
                row1 = worksheet.row(new_curr_row)
                if row1[0].value != ' ':
                    if float(row1[0].value) > 0:
                        prod_cate = row1[1].value
                        prod_name = str(int(row1[2].value))
                        description=row1[3].value
                        quantity = float(row1[4].value)
                        sale_price = float(row1[5].value)
                        product_id = self.env['product.product'].search([('name','=',prod_name)])
                        if not product_id:
                            product_id = self.env['product.product'].create([{'name':prod_name}])
                        order_line.append((0,0,{'product_id':product_id.id,
                                            'name':description,
                                            'product_uom_qty':quantity,
                                            'price_unit':sale_price,
                                            'order_id':self.id
                                            }))
                new_curr_row = new_curr_row + 1
            new_order_id.order_line = order_line
            new_order_id.action_confirm()
            new_order_id._create_invoices()
            for invoice in new_order_id.invoice_ids:
                invoice.write({'invoice_date':new_order_id.aloha_date})
                invoice.action_post()
                self.get_invoice_payment(dummy_aloha_inv_file_path,xlsx_dir_path,invoice)
           
    def get_invoice_payment(self,dummy_aloha_inv_file_path,xlsx_dir_path,invoice):
        import datetime
        pre_date = datetime.datetime.today() - datetime.timedelta(days=1)
        today_date = pre_date.strftime("%Y%m%d")

        csv_data = []
        with open(dummy_aloha_inv_file_path) as file_obj:
            reader = csv.reader(file_obj)
            for row in reader:
                csv_data.append(row)
        real_xlsx_path = xlsx_dir_path+'/'+today_date+'_inv'+'.xlsx'
        workbook = xlsxwriter.Workbook(real_xlsx_path)
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        for row in csv_data:
            sheet.append(row)
        workbook.save(real_xlsx_path)

        workbook = xlrd.open_workbook(real_xlsx_path)
        worksheet = workbook.sheet_by_name(workbook.sheet_names()[0])
        num_rows = worksheet.nrows
        curr_row = 7
        missed_list = []
        updated_num_rows = 0
        while curr_row < num_rows:
            row1 = worksheet.row(curr_row)
            
            try:
                if row1[0].value == "********* SUMMARY *********":
                    updated_num_rows = curr_row
                    break
            except IOError:
                print("Formate Error")
            curr_row = curr_row + 1
        new_curr_row = 7
        order_line = []
        while new_curr_row < updated_num_rows:
            row1 = worksheet.row(new_curr_row)
            if row1[0].value != ' ' and row1[0].value not in ["Total NAF:","Check #","Total Maestro:","-------------","Total VISA:"]:
                account_jr_obj = self.env['account.journal']
                account_jr_id = account_jr_obj.search([('type','=','cash')],limit=1)
                if row1[0].value == "*********  NAF   *********":
                    jr_id = account_jr_id.search([('type','=','cash')],limit=1)
                    if jr_id:
                        acc_jr_id = jr_id
                    else:
                        acc_jr_id = account_jr_id
                if row1[0].value == "*********  Maestro   *********":
                    jr_id = account_jr_id.search([('type','=','bank'),('code','=','Mastr')],limit=1)
                    if jr_id:
                        acc_jr_id = jr_id
                    else:
                        acc_jr_id = account_jr_id
                if row1[0].value == "*********  VISA   *********":
                    jr_id = account_jr_id.search([('code','=','visa'),('type','=','bank')],limit=1)
                    if jr_id:
                        acc_jr_id = jr_id
                    else:
                        acc_jr_id = account_jr_id
                
                if row1[0].value != '' and row1[0].value not in ['*********  NAF   *********','*********  Maestro   *********','*********  VISA   *********','*********  Maestro   *********','*********  VISA   *********']:
                    if float(row1[0].value) > 0:

                        payment = self.env['account.payment'].create({'payment_type' :'inbound',
                            'partner_id' : invoice.partner_id.id,
                            'amount' : float(row1[4].value),
                            "ref" : invoice.name,
                            "journal_id":acc_jr_id.id})
                        payment.action_post()
                        move_lines = payment.line_ids.filtered(lambda line: line.account_internal_type in ('receivable', 'payable') and not line.reconciled)
                        for line in move_lines:
                            invoice.js_assign_outstanding_line(line.id)
                    
            new_curr_row = new_curr_row + 1



